import logging
import os
import pandas as pd
from datetime import datetime
from config import COMPLETE_FOLDER
from database import get_data_by_inventory_date
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    filename="extract_data.log",
    filemode="a",
)
logger = logging.getLogger(__name__)

# 콘솔에 로그 출력을 위한 핸들러 추가
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)


def get_user_input() -> tuple:
    while True:
        try:
            start_date = input("시작 날짜를 입력하세요 (YYYY-MM-DD): ")
            end_date = input("종료 날짜를 입력하세요 (YYYY-MM-DD): ")
            datetime.strptime(start_date, "%Y-%m-%d")
            datetime.strptime(end_date, "%Y-%m-%d")
            logger.info(f"입력된 날짜: {start_date}부터 {end_date}까지")
            return start_date, end_date
        except ValueError:
            logger.error(
                "잘못된 날짜 형식입니다. YYYY-MM-DD 형식으로 다시 입력해주세요."
            )
            print("잘못된 날짜 형식입니다. YYYY-MM-DD 형식으로 다시 입력해주세요.")


def preprocess_extracted_data(df: pd.DataFrame) -> pd.DataFrame:
    logger.info("추출된 데이터 전처리 시작")
    df.columns = df.columns.str.replace(" ", "_").str.lower()

    # putawaydate 처리
    if "putawaydate" in df.columns:
        df["putawaydate"] = pd.to_datetime(df["putawaydate"], errors="coerce")

    # inventorydate 처리
    if "inventorydate" in df.columns:
        df["inventorydate"] = pd.to_datetime(df["inventorydate"]).dt.date
        # inventorydate를 'YY-MM-DD' 형식의 문자열로 변환
        df["inventorydate_str"] = df["inventorydate"].apply(lambda x: x.strftime("%y-%m-%d") if pd.notnull(x) else None)

    # count_po 처리: null이 아닌 값을 1로, null인 값을 0으로 변경
    if "count_po" in df.columns:
        df["count_po"] = df["count_po"].notnull().astype(int)

    # quantity 처리
    if "quantity" in df.columns:
        df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(0)

    # InventoryDate 기준으로 오름차순 정렬
    df = df.sort_values(by="inventorydate")

    logger.info("데이터 전처리 완료")
    return df


def analyze_data(df: pd.DataFrame) -> dict:
    # Weekly Analysis
    weekly_analysis = (
        df.groupby("week")
        .agg(
            {
                "count_po": "sum",
                "quantity": "sum",
            }
        )
        .reset_index()
    )
    weekly_analysis = weekly_analysis.sort_values("week")

    # ShipTo Analysis
    shipto_analysis = (
        df.groupby("shiptocode")
        .agg(
            {
                "count_po": "sum",
                "quantity": "sum",
            }
        )
        .reset_index()
    )

    total_count_po = shipto_analysis["count_po"].sum()
    total_quantity = shipto_analysis["quantity"].sum()

    shipto_analysis["count_po_ratio"] = shipto_analysis["count_po"] / total_count_po
    shipto_analysis["quantity_ratio"] = shipto_analysis["quantity"] / total_quantity

    shipto_analysis = shipto_analysis.sort_values("count_po", ascending=False)

    # Total 행 추가
    total_row = pd.DataFrame(
        {
            "shiptocode": ["TOTAL"],
            "count_po": [total_count_po],
            "quantity": [total_quantity],
            "count_po_ratio": [1.0],
            "quantity_ratio": [1.0],
        }
    )
    shipto_analysis = pd.concat([shipto_analysis, total_row])

    logger.info("주차별 및 ShipTo 데이터 분석 완료")
    return {"weekly_analysis": weekly_analysis, "shipto_analysis": shipto_analysis}


def adjust_column_width(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width


def save_to_excel(data: dict, filename: str):
    file_path = os.path.join(COMPLETE_FOLDER, filename)

    full_data_columns = [
        "receiptno",
        "replen_balance_order",
        "cust_sys_no",
        "allocated_part",
        "edi_order_type",
        "shipfromcode",
        "shiptocode",
        "country",
        "quantity",
        "putawaydate",
        "inventorydate",
        "fy",
        "quarter",
        "month",
        "week",
        "ordertype",
        "count_po",
    ]

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        logger.info("전체 데이터를 Excel 파일로 저장 중")
        # InventoryDate로 정렬된 데이터 저장
        sorted_full_data = data["full_data"].sort_values(by="inventorydate")

        # inventorydate_str 열을 사용하여 저장
        output_data = sorted_full_data[full_data_columns].copy()
        output_data.rename(columns={"inventorydate_str": "inventorydate"}, inplace=True)
        output_data.to_excel(writer, sheet_name="Full_Data", index=False)

        data["full_data"][full_data_columns].to_excel(
            writer, sheet_name="Full_Data", index=False
        )

        logger.info("Summary 데이터를 Excel 파일로 저장 중")
        summary_sheet = writer.book.create_sheet("Summary")

        # ShipTo 데이터 추가
        headers = [
            "ShipTo Code",
            "Count PO",
            "Quantity",
            "Count PO Ratio",
            "Quantity Ratio",
        ]
        for col, header in enumerate(headers, start=1):
            summary_sheet.cell(row=2, column=col, value=header)

        shipto_data = data["shipto_analysis"]
        for row, (_, data_row) in enumerate(shipto_data.iterrows(), start=3):
            summary_sheet.cell(row=row, column=1, value=data_row["shiptocode"])
            summary_sheet.cell(row=row, column=2, value=data_row["count_po"])
            summary_sheet.cell(row=row, column=3, value=data_row["quantity"])
            summary_sheet.cell(row=row, column=4, value=data_row["count_po_ratio"])
            summary_sheet.cell(row=row, column=5, value=data_row["quantity_ratio"])

        # 제목 추가 및 병합
        last_row = summary_sheet.max_row
        summary_sheet.merge_cells(f"A1:E1")
        summary_sheet["A1"] = "Weekly Analysis"
        summary_sheet["A1"].alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # Week 정보 추가
        week_cell = summary_sheet.cell(row=2, column=7, value="Week:")
        week_value_cell = summary_sheet.cell(
            row=2, column=8, value=data["weekly_analysis"]["week"].iloc[0]
        )
        week_cell.font = Font(bold=True)

    workbook = load_workbook(file_path)

    # Full Data 시트 조정
    full_data_sheet = workbook["Full_Data"]
    adjust_column_width(full_data_sheet)

    summary_sheet = workbook["Summary"]

    # 스타일 정의
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )
    light_blue_fill = PatternFill(
        start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
    )

    # 제목 스타일 적용
    summary_sheet["A1"].font = title_font

    # 헤더 스타일 적용
    for cell in summary_sheet[2][:5]:  # Quantity Ratio까지만 스타일 적용
        cell.font = header_font
        cell.border = border
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ShipTo Code 행 (헤더) 스타일 적용
    for cell in summary_sheet[2][:5]:
        cell.fill = light_blue_fill

    # 데이터 셀에 테두리 적용 및 비율 포맷 설정
    for row in summary_sheet.iter_rows(
        min_row=3, max_row=summary_sheet.max_row, min_col=1, max_col=5
    ):
        for cell in row:
            cell.border = border
            if cell.column in [4, 5]:  # Count PO Ratio와 Quantity Ratio 열
                cell.number_format = "0.00%"

        # TOTAL 행에 옅은 파란색 배경 적용
        if row[0].value == "TOTAL":
            for cell in row:
                cell.fill = light_blue_fill

    # Summary 시트의 열 너비 자동 조정 (A부터 E열까지만)
    for column_letter in ["A", "B", "C", "D", "E"]:
        max_length = 0
        column = summary_sheet[column_letter]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        summary_sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(file_path)
    logger.info(f"데이터가 Excel 파일로 저장되었습니다: {file_path}")


def main():
    start_date, end_date = get_user_input()

    start_date = datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.strptime(end_date, "%Y-%m-%d")
    logger.info(f"날짜 변환 완료: {start_date}부터 {end_date}까지")

    logger.info(
        f"데이터베이스에서 {start_date.date()}부터 {end_date.date()}까지의 데이터를 추출 중"
    )
    df = get_data_by_inventory_date(start_date, end_date)
    logger.info(f"데이터베이스에서 {len(df)}개의 데이터를 추출 완료")

    if not df.empty:
        df = preprocess_extracted_data(df)
        analysis_results = analyze_data(df)

        filename = f"{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}_ReceivingTAT_analysis.xlsx"

        data_to_save = {
            "full_data": df,
            "weekly_analysis": analysis_results["weekly_analysis"],
            "shipto_analysis": analysis_results["shipto_analysis"],
        }

        logger.info("Excel 파일 생성 시작")
        save_to_excel(data_to_save, filename)
        logger.info("Excel 파일 생성 완료")
    else:
        logger.info("지정된 날짜 범위에 해당하는 데이터가 없습니다")

    logger.info("프로그램 종료")


if __name__ == "__main__":
    main()
